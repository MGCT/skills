#!/usr/bin/env python3
"""Build a styled Excel timeline from a JSON spec, with a month-grid Gantt.

The model does the thinking — extracting milestones from the project and
confirming dates with the user — and hands this script a clean JSON spec. The
script does the deterministic part: a formatted schedule table plus a
month-by-month Gantt grid coloured by status, and a list of anything still
unscheduled (TBC). Keeps date/openpyxl mechanics out of the model's hands.

Spec (JSON):
{
  "title": "Northwind — Delivery Timeline",
  "status_as_of": "2026-06-29",
  "items": [
    {"workstream": "Data foundation", "milestone": "Single searchable store",
     "owner": "Acme", "start": "2026-04", "end": "2026-09",
     "status": "in progress", "depends_on": "", "assumption": ""},
    {"workstream": "Tooling", "milestone": "UK pilot", "status": "planned",
     "assumption": "date TBC — not agreed with client"}   // no start/end -> Unscheduled
  ]
}

Dates: "YYYY-MM" or "YYYY-MM-DD" (month granularity). Items without a valid
start AND end go to an "Unscheduled / TBC" block instead of the Gantt.
Status: done | in progress | planned | at risk (others render grey).

Usage:  python build_xlsx.py SPEC.json OUTPUT.xlsx
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

STATUS_FILL = {
    "done": "2D6A4F",
    "in progress": "E8A317",
    "planned": "9AA6B2",
    "at risk": "B23A48",
}
ACCENT = "2D6A4F"
ACCENT_SOFT = "E8F0EC"
INK = "1A1F2B"
RULE = "DFE3E8"

THIN = Side(style="thin", color=RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_ym(s):
    """Return (year, month) from 'YYYY-MM' or 'YYYY-MM-DD', else None."""
    if not s or not isinstance(s, str):
        return None
    parts = s.strip().split("-")
    try:
        y, m = int(parts[0]), int(parts[1])
        if 1 <= m <= 12 and 1900 < y < 2100:
            return (y, m)
    except (ValueError, IndexError):
        pass
    return None


def months_between(a, b):
    """Inclusive list of (year, month) from a to b."""
    out, (y, m) = [], a
    while (y, m) <= b:
        out.append((y, m))
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return out


def ym_label(ym):
    names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{names[ym[1] - 1]} {str(ym[0])[2:]}"


def build(spec, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"

    title = spec.get("title", "Timeline")
    as_of = spec.get("status_as_of", "")
    items = spec.get("items", [])

    white = Font(color="FFFFFF", bold=True, size=14)
    bold = Font(bold=True, color=INK)
    muted = Font(color="5B6472", italic=True, size=9)
    head_fill = PatternFill("solid", fgColor=ACCENT)
    soft_fill = PatternFill("solid", fgColor=ACCENT_SOFT)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title band
    ws["A1"] = title
    ws["A1"].font = white
    ws["A1"].fill = head_fill
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    if as_of:
        ws["A2"] = f"Status as of {as_of}"
        ws["A2"].font = muted

    # Split scheduled vs unscheduled
    scheduled = []
    for it in items:
        s, e = parse_ym(it.get("start")), parse_ym(it.get("end"))
        if s and e and s <= e:
            scheduled.append((it, s, e))
    unscheduled = [it for it in items
                   if (it, parse_ym(it.get("start")), parse_ym(it.get("end")))
                   not in [(x[0], x[1], x[2]) for x in scheduled]]

    headers = ["Workstream", "Milestone", "Owner", "Start", "End", "Status"]
    start_row = 4
    grid_start_col = len(headers) + 1  # first Gantt month column

    # Month span across scheduled items
    span = []
    if scheduled:
        lo = min(s for _, s, _ in scheduled)
        hi = max(e for _, _, e in scheduled)
        span = months_between(lo, hi)

    # Header row
    r = start_row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.fill = soft_fill
        cell.border = BORDER
        cell.alignment = left if c <= 2 else center
    for c, ym in enumerate(span, grid_start_col):
        cell = ws.cell(r, c, ym_label(ym))
        cell.font = bold
        cell.fill = soft_fill
        cell.border = BORDER
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = 6

    # Data rows
    r += 1
    for it, s, e in scheduled:
        vals = [it.get("workstream", ""), it.get("milestone", ""),
                it.get("owner", ""), it.get("start", ""), it.get("end", ""),
                it.get("status", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = BORDER
            cell.alignment = left if c <= 2 else center
        fill_hex = STATUS_FILL.get(str(it.get("status", "")).lower(), "9AA6B2")
        bar = PatternFill("solid", fgColor=fill_hex)
        for c, ym in enumerate(span, grid_start_col):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if s <= ym <= e:
                cell.fill = bar
        dep = it.get("depends_on") or it.get("assumption")
        if dep:
            note = ws.cell(r, grid_start_col + len(span), f"⚠ {dep}" if it.get("assumption") else dep)
            note.font = muted
        r += 1

    # Unscheduled / TBC block
    if unscheduled:
        r += 1
        head = ws.cell(r, 1, "Unscheduled / TBC")
        head.font = Font(bold=True, color="B7791F")
        r += 1
        for c, h in enumerate(["Workstream", "Milestone", "Status", "Why unscheduled"], 1):
            cell = ws.cell(r, c, h)
            cell.font = bold
            cell.fill = soft_fill
            cell.border = BORDER
        r += 1
        for it in unscheduled:
            vals = [it.get("workstream", ""), it.get("milestone", ""),
                    it.get("status", ""),
                    it.get("assumption") or "no date set"]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.alignment = left
            r += 1

    # Legend
    r += 2
    ws.cell(r, 1, "Legend").font = bold
    r += 1
    for c, (label, hexv) in enumerate(STATUS_FILL.items(), 1):
        cell = ws.cell(r, c, label)
        cell.fill = PatternFill("solid", fgColor=hexv)
        cell.alignment = center
        cell.font = Font(color="FFFFFF" if label != "planned" else INK, bold=True, size=9)
        cell.border = BORDER

    # Column widths + freeze
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 13
    ws.freeze_panes = ws.cell(start_row + 1, grid_start_col)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(scheduled), len(unscheduled), len(span)


def main(argv=None):
    argv = argv or sys.argv[1:]
    if len(argv) < 2:
        sys.exit("Usage: python build_xlsx.py SPEC.json OUTPUT.xlsx")
    spec = json.loads(Path(argv[0]).read_text(encoding="utf-8"))
    sched, unsched, months = build(spec, argv[1])
    print(f"Wrote {argv[1]}  ({sched} scheduled across {months} months, "
          f"{unsched} unscheduled/TBC)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
